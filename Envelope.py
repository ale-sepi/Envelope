import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import plotly.graph_objs as go
import openpyxl

st.title('Envelopes Configurator')


def check_password():
    """Returns `True` if the user had the correct password."""

    def password_entered():
        """Checks whether a password entered by the user is correct."""
        if st.session_state["password"] == st.secrets["password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # don't store password
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # First run, show input for password.
        st.text_input(
            "Password", type="password", on_change=password_entered, key="password"
        )
        return False
    elif not st.session_state["password_correct"]:
        # Password not correct, show input + error.
        st.text_input(
            "Password", type="password", on_change=password_entered, key="password"
        )
        st.error("😕 Password incorrect")
        return False
    else:
        # Password correct.
        return True

if check_password():

    Name = st.text_input('Name', 'Insert name of the unit')
    dtmin = st.number_input('DT min [°C]')
    dtmax = st.number_input('DT max [°C]')


    df = pd.DataFrame(columns=['X [°C]','Y min [°C]','Y max [°C]'])
    config = {
        'X' : st.column_config.NumberColumn('X', required=True, width='large'),
        'ymin' : st.column_config.NumberColumn('ymin', required=True, width='large'),
        'ymax' : st.column_config.NumberColumn('ymax', required=True, width='large')
    }


    # Watch out, result is the new data frame!
    result = st.data_editor(df, column_config = config, num_rows='dynamic')


    # Extract data
    x_values = result['X [°C]'].astype(float)
    ymin_values = result['Y min [°C]'].astype(float)
    ymax_values = result['Y max [°C]'].astype(float)


    # Sorting with respect of x
    sorted_indices = sorted(range(len(x_values)), key=lambda k: x_values[k])
    x_values = [x_values[i] for i in sorted_indices]
    ymin_values = [ymin_values[i] for i in sorted_indices]
    ymax_values = [ymax_values[i] for i in sorted_indices]

    def Reverse(lst):
        new_lst = lst[::-1]
        return new_lst

    x_data = x_values + Reverse(x_values)
    y_data = ymin_values + Reverse(ymax_values)



    if len(x_data) > 0:
        x_data.append(x_data[0])
        y_data.append(y_data[0])
        # Create a Plotly trace for the polygon using a line plot
        polygon_trace = go.Scatter(
        x=x_data,
        y=y_data,
        mode="lines+markers",
        marker=dict(size=10, color='blue'),
        fill="toself",  # Filled polygon
        fillcolor="rgba(0,100,80,0.2)",  # Polygon fill color with opacity
        line=dict(color="blue"),
    )

    # Create a Plotly layout
        layout = go.Layout(
        title=Name,
        titlefont = dict(size=22),
        xaxis=dict(title='ELWT [°C]', autorange=True, showgrid=True),  # Customize x-axis range if needed
        yaxis=dict(title= 'OAT [°C]', autorange=True, showgrid=True),
        
    )

        # Create a Plotly figure
        fig = go.Figure(data=[polygon_trace], layout=layout)

        # Render the Plotly figure in the Streamlit app
        st.plotly_chart(fig)




    # Check if a new row was added
    if result is not None and 'new_row' in result:
        new_row = pd.DataFrame([result['new_row']])
        df = pd.concat([df, new_row], ignore_index=True)

        st.plotly_chart(fig)


    # Excel function to write in the correct format the envelop, ready to be uploaded
    # on dds
    def write_excel_file(data1, data2, data3):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        sheet['D1'] = "Coefficients"
        sheet['A3'] = Name
        sheet['B3'] = dtmax
        sheet['C3'] = dtmin
        headers = ["Name", "dtmax", "dtmin", "X", "Y_min", "Y_Max"]
        
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
        
        for index, value in enumerate(data1):
            cell = sheet.cell(row=index + 3, column=4)  # D3 corresponds to row 3, column 4
            cell.value = round(value,2)

        for index, value in enumerate(data2):
            cell = sheet.cell(row=index + 3, column=5)  # E3 corresponds to row 3, column 5
            cell.value = round(value,2)

        for index, value in enumerate(data3):
            cell = sheet.cell(row=index + 3, column=6)  # F3 corresponds to row 3, column 6
            cell.value = round(value,2)

        return workbook
    


    def download_excel():
        workbook = write_excel_file(x_values, ymin_values, ymax_values)
        file_name = Name + ".xlsx"
        workbook.save(file_name)

        with open(Name + ".xlsx", 'rb') as f:
            data = f.read()
            st.download_button(label='Download Excel File', data=data, file_name=Name + '.xlsx', key='download_button', mime='application/vnd.ms-excel')

    

    download_excel()
